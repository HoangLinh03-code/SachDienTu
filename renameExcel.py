import os
from openpyxl import load_workbook

# Đường dẫn tới folder gốc
root_folder = r"C:\Users\Admin\Desktop\Maru\SachDienTu\SDT_TOAN\SDT_TOAN_SGK đã fix"

# Duyệt toàn bộ cây thư mục
for root, dirs, files in os.walk(root_folder):
    for file in files:
        if file.lower().endswith(".xlsx"):
            file_path = os.path.join(root, file)
            file_name = os.path.splitext(file)[0]
            print(f"🔹 Đang xử lý: {file_path}")

            try:
                wb = load_workbook(file_path)
                # Kiểm tra xem sheet "Cay Kien Thuc" có tồn tại không
                if "Cay Kien Thuc" in wb.sheetnames:
                    sheet = wb["Cay Kien Thuc"]

                    for row in sheet.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str):
                                text = cell.value
                                # Nếu chứa TAP1
                                if "TAP1" in text:
                                    text = text.replace(file_name, f"{file_name}_1")
                                    text = text.replace("TAP1", "")
                                # Nếu chứa TAP2
                                elif "TAP2" in text:
                                    text = text.replace(file_name, f"{file_name}_2")
                                    text = text.replace("TAP2", "")
                                # Ghi lại nếu có thay đổi
                                if text != cell.value:
                                    cell.value = text

                    wb.save(file_path)
                    print(f"✅ Đã cập nhật sheet 'Cay Kien Thuc' trong {file}")
                else:
                    print(f"⚠️ Bỏ qua {file} — không có sheet 'Cay Kien Thuc'")

                wb.close()
            except Exception as e:
                print(f"❌ Lỗi khi xử lý {file}: {e}")
