import os
import openpyxl

def merge_excels(parent_folder, output_file):
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    row_out = 1

    for root, dirs, files in os.walk(parent_folder):
        for file in files:
            if file.endswith(".xlsx"):
                file_path = os.path.join(root, file)
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if not any(row):
                        continue
                    row_list = list(row)
                    # Thêm ô rỗng theo số lượng cột
                    if len(row_list) == 1:
                        row_list = ["", ""] + row_list
                    elif len(row_list) == 2:
                        row_list = [""] + row_list
                    # Nếu đủ 3 cột thì giữ nguyên
                    ws_out.append(row_list)
                    row_out += 1
                # Thêm một dòng trống sau khi thêm dữ liệu từ một file nhỏ
                ws_out.append([])
                row_out += 1

    wb_out.save(output_file)
    print(f"Đã tạo file tổng hợp: {output_file}")

if __name__ == "__main__":
    parent_folder = r"C:\Users\Admin\Desktop\Maru\SachDienTu\SDT_TOAN\SDT_TOAN_SGK đã fix"
    output_file = os.path.join(parent_folder, "TongHopCTST.xlsx")
    merge_excels(parent_folder, output_file)