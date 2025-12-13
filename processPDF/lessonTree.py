import json, os
import openpyxl

def process_lesson_tree(pdf_path, json_path, output_folder):
    file_name = os.path.basename(pdf_path).replace(".pdf", "")
    if not os.path.exists(os.path.join(output_folder, file_name)):
        os.makedirs(os.path.join(output_folder, file_name))

    with open(json_path, "r", encoding="utf-8") as f:
        bookDatas = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cay Kien Thuc"

    lessons_json = []

    def write_tree(data, row, col, parent_id=""):
        for item in data:
            Lid = item.get("Lid", "")
            # Sinh id theo d·∫°ng cha_con
            if parent_id:
                cur_id = f"{parent_id}_{Lid}"
            else:
                cur_id = f"{file_name}_{Lid}"
            Name = item.get("Name", "")
            excel_name = f"\"{cur_id}\":\"{Name}\""
            ws.cell(row=row, column=col, value=excel_name)
            next_row = row + 1
            # N·∫øu c√≥ Content l√† list, duy·ªát ti·∫øp
            if "Content" in item and isinstance(item["Content"], list) and item["Content"]:
                for lesson in item["Content"]:
                    # N·∫øu lesson l√† c·∫•p cu·ªëi (c√≥ St, End)
                    if "St" in lesson and "End" in lesson:
                        lesson_Lid = lesson.get("Lid", "")
                        lesson_id = f"{cur_id}_{lesson_Lid}"
                        lesson_Name = lesson.get("Name", "")
                        lesson_excel_name = f"\"{lesson_id}\":\"{lesson_Name}\""
                        ws.cell(row=next_row, column=col+1, value=lesson_excel_name)
                        # Th√™m v√†o json
                        lessons_json.append({
                            "Name": lesson_Name,
                            "Lid": lesson_id,
                            "St": lesson.get("St", ""),
                            "End": lesson.get("End", "")
                        })
                        next_row += 1
                    # N·∫øu lesson c√≤n l·ªìng Content, ƒë·ªá quy
                    elif "Content" in lesson and isinstance(lesson["Content"], list):
                        next_row = write_tree([lesson], next_row, col+1, cur_id)
                    else:
                        # Tr∆∞·ªùng h·ª£p lesson kh√¥ng c√≥ St/End v√† kh√¥ng c√≥ Content
                        lesson_Lid = lesson.get("Lid", "")
                        lesson_id = f"{cur_id}_{lesson_Lid}"
                        lesson_Name = lesson.get("Name", "")
                        lesson_excel_name = f"\"{lesson_id}\":\"{lesson_Name}\""
                        ws.cell(row=next_row, column=col+1, value=lesson_excel_name)
                        next_row += 1
                row = next_row
            else:
                # N·∫øu l√† c·∫•p cu·ªëi c√πng v√† c√≥ St, End
                if "St" in item and "End" in item:
                    lessons_json.append({
                        "Name": item.get("Name", ""),
                        "Lid": cur_id,
                        "St": item.get("St", ""),
                        "End": item.get("End", "")
                    })
                row += 1
        return row

    write_tree(bookDatas, 1, 1)

    excel_output = os.path.join(output_folder, file_name, f"{file_name}.xlsx")
    wb.save(excel_output)
    print(f"ƒê√£ t·∫°o file Excel: {excel_output}")

    # Ghi file json ch·ª©a c√°c b√†i h·ªçc cu·ªëi c√πng
    json_output = os.path.join(output_folder, file_name, f"{file_name}.json")
    with open(json_output, "w", encoding="utf-8") as f:
        json.dump(lessons_json, f, ensure_ascii=False, indent=2)
    print(f"ƒê√£ t·∫°o file JSON: {json_output}")

def scan_folder(folder, output_folder):
    for root, dirs, files in os.walk(folder):
        for f in files:
            if f.lower().endswith(".pdf"):
                pdf_path = os.path.join(root, f)
                json_path = os.path.join(root, f.replace(".pdf", ".json"))
                process_lesson_tree(pdf_path, json_path, output_folder)

if __name__ == "__main__":
    # listL = [
    #     "SDT_NGUVAN"
    # ]
    # folder_path = r"D:\\pdf\\SDT_NGUVAN_KNTT_C11"
    # for item in listL:
    #     folder = os.path.join(folder_path, item, f"{item}_SGK")
    #     for root, dirs, files in os.walk(folder):
    #         for d in dirs:
    #             class_folder = os.path.join(root, d)
    #             output_folder = os.path.join(folder_path, item, f"{item}_SGK ƒë√£ fix", d)
    #             scan_folder(class_folder, output_folder)
    # ƒê∆∞·ªùng d·∫´n ƒë·∫øn folder Ng·ªØ VƒÉn KNTT 11 c·ªßa b·∫°n
    folder_path = r"D:\\pdf\\SDT_NGUVAN_KNTT_C11" 

    # L∆∞u file Excel v√† JSON t·ªïng ngay t·∫°i ƒë√≥
    output_folder = folder_path 

    print(f"üìÇ ƒêang qu√©t: {folder_path}")
    scan_folder(folder_path, output_folder)