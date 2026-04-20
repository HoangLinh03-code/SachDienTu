import json
import os
import openpyxl
import re
def create_excel_like_sample(json_path, book_code, book_name, output_path):
    if not os.path.exists(json_path):
        print(f"❌ Không tìm thấy file JSON: {json_path}")
        return

    print(f"🎨 Đang tạo Excel theo mẫu 'bậc thang' cho: {book_code}...")
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # =========================================================================
    # --- [NEW] TỰ ĐỘNG NHẬN DIỆN "TẬP" ĐỂ SỬA ROOT ID (Lid) ---
    # =========================================================================
    file_name = os.path.basename(json_path)
    # Tìm các mẫu như: Tap 2, Tập hai, Tập 3...
    match = re.search(r't[aâậ]p[\s_\-]*(m[ộo]t|hai|ba|b[ốo]n|\d+)', file_name, re.IGNORECASE)
    
    if match:
        val = match.group(1).lower()
        # Chuyển đổi chữ thành số nếu cần
        mapping = {'một': '1', 'mot': '1', 'hai': '2', 'ba': '3', 'bốn': '4', 'bon': '4'}
        tap_number = mapping.get(val, val)
        print(f"   ⚠️ Phát hiện Sách TẬP {tap_number} -> Đang ép Root ID thành '{tap_number}'...")
    else:
        tap_number = "1"
        print(f"   ℹ️ Không phát hiện 'Tập X' trong tên, mặc định Root ID = '1'.")

    # Cập nhật Lid của Node Gốc để ép Excel đổi mã
    if isinstance(data, list) and len(data) > 0:
        data[0]["Lid"] = str(tap_number)
    elif isinstance(data, dict):
        data["Lid"] = str(tap_number)
    # =========================================================================

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cay Kien Thuc"
    
    # Tắt đường lưới để nhìn giống mẫu hơn (tuỳ chọn)
    ws.sheet_view.showGridLines = False 

    # --- GHI DÒNG ĐẦU TIÊN (TÊN SÁCH - CẤP 0) ---
    root_value = f"\"{book_code}\":\"{book_name}\""
    ws.cell(row=1, column=1, value=root_value)

    current_row = 2

    # Hàm đệ quy để ghi dữ liệu
    def write_node(node_list, level, parent_id=""):
        nonlocal current_row
        for item in node_list:
            Lid = item.get("Lid", "")
            Name = item.get("Name", "")

            # Tạo ID dạng chuỗi (VD: "SDT_..._2", "SDT_..._2_1")
            if parent_id:
                short_id = f"{parent_id}_{Lid}"
            else:
                short_id = f"{book_code}_{Lid}"

            cell_value = f"\"{short_id}\":\"{Name}\""
            ws.cell(row=current_row, column=level, value=cell_value)
            
            current_row += 1

            if "Content" in item and isinstance(item["Content"], list):
                write_node(item["Content"], level + 1, short_id)

    # Bắt đầu duyệt từ dữ liệu JSON (Level 2 = Cột B)
    if isinstance(data, list):
        write_node(data, 2)
    elif isinstance(data, dict):
        write_node([data], 2)

    # Tự động chỉnh độ rộng cột cho dễ nhìn
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 5

    wb.save(output_path)
    print(f"✅ Đã xuất file chuẩn mẫu: {output_path}")

if __name__ == "__main__":
    # ================= CẤU HÌNH =================
    
    # 1. Thư mục làm việc
    working_dir = r"D:\NguVan\C12_CTST"
    
    # 2. File JSON đầu vào (File SGK chuẩn)
    json_input = os.path.join(working_dir, "SHS NGU VAN 12 TAP 2 CTST (Ruot ITB 06.02.25).json")
    
    # 3. Mã sách (ROOT KEY) - Sửa lại cho đúng mã dự án của bạn
    # Trong ảnh mẫu là: SDT_NGUVAN_KNTT_C11 (Không có _1 ở cuối)
    my_book_code = "SDT_NGUVAN_CTST_C12" 
    
    # 4. Tên sách (ROOT NAME) - Hiển thị ở dòng đầu tiên
    my_book_name = "Ngữ văn lớp 12 bộ Chân trời sáng tạo"
    
    # 5. Tên file Excel đầu ra
    excel_output = os.path.join(working_dir, "CayKienThuc_CTST_C12.xlsx")

    # ============================================
    
    create_excel_like_sample(json_input, my_book_code, my_book_name, excel_output)